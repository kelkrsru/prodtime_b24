import datetime
import decimal
import logging

from rest_framework import viewsets
from django_filters.rest_framework import DjangoFilterBackend

import core.methods as core_methods

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist, BadRequest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from core.bitrix24.bitrix24 import (DealB24, ProductRowB24, SmartProcessB24, CompanyB24,
                                    ProductInCatalogB24, ListB24, create_portal, TemplateDocB24)
from core.models import Portals
from dealcard.serializers import ProdTimeDealSerializer
from settings.models import SettingsPortal, Numeric, AssociativeYearNumber
from dealcard.models import Deal, ProdTimeDeal

from pybitrix24 import Bitrix24

logger = logging.getLogger(__name__)
SEPARATOR = '*' * 40
NEW_STR = '\n    '


@xframe_options_exempt
@csrf_exempt
def index(request):
    """Метод страницы Карточка сделки."""
    template: str = 'dealcard/index.html'
    title: str = 'Страница карточки сделки'

    logger.info(f'{SEPARATOR}')
    logger.info(f'Запущено приложение Срок производства')
    try:
        member_id, deal_id, auth_id = core_methods.initial_check(request)
        logger.info(f'Полученные параметры из запроса: {member_id=}, {deal_id=}')
    except BadRequest:
        logger.error(f'Ошибка запроса: неизвестный тип запроса - {request.method=}')
        return render(request, 'error.html', {'error_name': 'QueryError',
                                              'error_description': 'Неизвестный тип запроса'})
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)
    user_info = core_methods.get_current_user(request, auth_id, portal, settings_portal.is_admin_code)
    year = int(datetime.date.today().year)
    Numeric.objects.get_or_create(portal=portal, year=year, defaults={'last_number': 0})

    try:
        deal_bx = DealB24(portal, deal_id)
        deal_bx.get_all_products()
        logger.debug(f'Полученные товарные позиции сделки:{NEW_STR}{deal_bx.products}')

        deal, created = Deal.objects.get_or_create(portal=portal, deal_id=deal_bx.id,
                                                   defaults={'general_number': '000.'})
        logger.info(f'Сделка: {created=} {deal.deal_id=}')
        deal.responsible = core_methods.get_responsible_deal(portal, deal_bx)
        logger.info(f'Ответственный за сделку: {deal.responsible.id_b24=} {deal.responsible.last_name} '
                    f'{deal.responsible.first_name}')
        deal.invoice_number = deal_bx.properties.get(settings_portal.num_invoice_code)
        logger.info(f'Номер счета для сделки: {deal.invoice_number=}')
        deal.save()
    except Exception as ex:
        logger.error(f'Ошибка при получении или создании сделки: Ошибка={ex.args[0]}, Описание={ex.args[1]}')
        return render(request, 'error.html', {'error_name': ex.args[0], 'error_description': ex.args[1]})

    logger.info(f'Производим перебор товарных позиций из сделки битрикс24')
    for product in deal_bx.products:
        logger.info(f'Товарная позиция: {product.get("PRODUCT_ID")=}')
        # Проверка, что товар является товаром каталога
        try:
            if not product['PRODUCT_ID'] or int(product['PRODUCT_ID']) <= 0:
                raise RuntimeError('Error', 'Product not found')
            product_in_catalog = ProductInCatalogB24(portal, int(product['PRODUCT_ID']))
            logger.debug(f'Товар является товаром каталога:{NEW_STR}{product_in_catalog.properties=}')
        except RuntimeError as ex:
            logger.error(f'Товар ID={product.get("ID")} Name={product.get("PRODUCT_NAME")} не найден в каталоге товаров'
                         f' Битрикс24')
            return render(request, 'error.html', {
                'error_name': ex.args[1],
                'error_description': f'Товар {product["PRODUCT_NAME"]} не найден в каталоге товаров Битрикс24'
            })

        product_value: dict = dict()
        for key, value in product.items():
            if key in core_methods.codes_values:
                if core_methods.codes_values.get(key)[1] == 'int':
                    new_value = int(value)
                elif core_methods.codes_values.get(key)[1] == 'decimal':
                    new_value = round(decimal.Decimal(value), 2) or 0
                elif core_methods.codes_values.get(key)[1] == 'bool':
                    new_value = True if value == 'Y' else False
                else:
                    new_value = str(value)
                product_value[core_methods.codes_values.get(key)[0]] = new_value
        product_value['tax_sum'] = round(product_value.get('quantity') * product_value.get('price_exclusive') *
                                         product_value.get('tax') / 100, 2)
        product_value['sum'] = round(product_value.get('quantity') * product_value.get('price_account'), 2)
        product_value['portal'] = portal
        logger.debug(f'Свойства товарной позиции для обновления или создания:{NEW_STR}{product_value=}')

        prodtime, created = ProdTimeDeal.objects.update_or_create(product_id_b24=product_value.get('product_id_b24'),
                                                                  defaults=product_value)
        if created:
            prodtime.name_for_print = prodtime.name
        logger.info(f'Товарная позиция: {created=} {prodtime.id=}')

        logger.info(f'Работа с прямыми затратами, нормочасами, материалами, сроком производства из каталога')
        # Работа с прямыми затратами, нормочасами, материалами, сроком производства из каталога
        core_methods.set_fields_from_catalog_b24(prodtime, settings_portal, product_in_catalog)

        logger.info(f'Работа с прибылью')
        # Работа с прибылью
        logger.info(f'Свойство is_change_income имеет значение {prodtime.is_change_income}')
        if (deal_bx.properties.get(settings_portal.deal_field_code_income_res) and not prodtime.income
                and not prodtime.is_change_income):
            core_methods.calculation_income(prodtime, settings_portal)

        logger.info(f'Работа с эквивалентом')
        # Работа с эквивалентом
        logger.info(f'Свойство is_change_equivalent имеет значение {prodtime.is_change_equivalent}')
        if prodtime.is_change_equivalent:
            product_value['equivalent'] = str(prodtime.equivalent).replace(',', '.')
        else:
            try:
                product_in_catalog = ProductB24Old(portal, prodtime.product_id_b24)
            except RuntimeError as ex:
                return render(request, 'error.html', {
                    'error_name': ex.args[0],
                    'error_description': f'{ex.args[1]} для товара {prodtime.name}'
                })
            equivalent_code = settings_portal.equivalent_code
            if equivalent_code not in product_in_catalog.props or not product_in_catalog.props.get(equivalent_code):
                product_value['equivalent'] = ''
                prodtime.equivalent = 0
            else:
                product_value['equivalent'] = list(product_in_catalog.props.get(equivalent_code).values())[1]
                prodtime.equivalent = decimal.Decimal(product_value['equivalent'])
            logger.info(f'Для свойства equivalent установлено значение {prodtime.equivalent}')
            prodtime.save()
        if prodtime.equivalent and prodtime.quantity:
            prodtime.equivalent_count = (prodtime.equivalent * prodtime.quantity)
        else:
            prodtime.equivalent_count = 0
        logger.info(f'Для свойства equivalent_count установлено значение {prodtime.equivalent_count}')
        prodtime.save()

    logger.info(f'Удаление лишних товарных позиций, которые удалены в сделке')
    # Удаление лишних товарных позиций, которые удалены в сделке
    products_in_db = ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id)
    for product in products_in_db:
        if not next((x for x in deal_bx.products if int(x.get('ID')) == product.product_id_b24), None):
            logger.info(f'{product.id} удален из БД, так как отсутствует в сделке')
            product.delete()

    # Подсчет суммарного эквивалента
    sum_equivalent = core_methods.count_sum_equivalent(ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id))
    logger.info(f'Суммарный эквивалент равен {sum_equivalent=}')

    products_in_db.order_by('sort')

    max_prodtime = core_methods.count_set_max_prodtime(member_id=member_id, deal_id=deal_id)
    logger.info(f'Максимальный срок производства равен {max_prodtime=}')
    # max_prodtime = deal_bx.properties.get(settings_portal.max_prodtime_code)
    # if max_prodtime:
    #     max_prodtime = datetime.datetime.strptime(max_prodtime.split('T')[0], '%Y-%m-%d').date()
    # else:
    #     max_prodtime = 'Не задан'

    context = {
        'title': title,
        'products': products_in_db,
        'sum_equivalent': sum_equivalent,
        'max_prodtime': max_prodtime,
        'member_id': member_id,
        'deal_id': deal_id,
        'deal': deal,
        'user': user_info
    }
    response = render(request, template, context)
    logger.info(f'{SEPARATOR}')
    if auth_id:
        response.set_cookie(key='user_id', value=user_info.get('user_id'))
    return response


@xframe_options_exempt
@csrf_exempt
def update_deal(request):
    """Метод вызова обновлений в БД (как в index) через webhook bitrix24."""

    logger.info(f'{SEPARATOR}')
    logger.info(f'Запущен метод обновления сделки через webhook')
    try:
        member_id, deal_id, token = core_methods.initial_check_with_token(request)
        logger.info(f'Полученные параметры из запроса: {member_id=}, {deal_id=}')
    except BadRequest:
        logger.error(f'Ошибка запроса: неизвестный тип запроса - {request.method=}')
        return HttpResponse(status=400, content=f'Error_name=BadRequest, Error_description=Неизвестный тип запроса')

    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)
    year = int(datetime.date.today().year)
    Numeric.objects.get_or_create(portal=portal, year=year, defaults={'last_number': 0})

    if not token or token != settings_portal.token:
        logger.error(f'Токен авторизации не передан или является неверным')
        return HttpResponse(status=401, content=f'Токен авторизации не передан или является неверным')

    try:
        deal_b24 = DealB24(portal, deal_id)
        deal_b24.get_all_products()
        logger.debug(f'Полученные товарные позиции сделки:{NEW_STR}{deal_b24.products}')

        responsible = core_methods.get_responsible_deal(portal, deal_b24)
        logger.info(f'Ответственный за сделку: {responsible.id_b24=} {responsible.last_name} {responsible.first_name}')

        deal, created = Deal.objects.get_or_create(
            portal=portal, deal_id=deal_b24.id,
            defaults={'general_number': '000.'}
        )
        logger.info(f'Сделка: {created=} {deal.deal_id=}')
        deal.responsible = responsible
        deal.invoice_number = deal_b24.properties.get(settings_portal.num_invoice_code)
        logger.info(f'Номер счета для сделки: {deal.invoice_number=}')
        deal.save()
    except Exception as ex:
        logger.error(f'Ошибка при получении или создании сделки: Ошибка={ex.args[0]}, Описание={ex.args[1]}')
        return HttpResponse(status=404, content=f'Error_name={ex.args[0]}, Error_description={ex.args[1]}')

    logger.info(f'Производим перебор товарных позиций из сделки битрикс24')
    for product in deal_b24.products:
        logger.info(f'Товарная позиция: {product.get("PRODUCT_ID")=}')
        # Проверка, что товар является товаром каталога
        try:
            if not product.get('PRODUCT_ID') or int(product.get('PRODUCT_ID')) <= 0:
                raise RuntimeError('Error', 'Product not found')
            product_in_catalog = ProductInCatalogB24(portal, int(product.get('PRODUCT_ID')))
            logger.debug(f'Товар является товаром каталога:{NEW_STR}{product_in_catalog.properties=}')
        except RuntimeError:
            logger.error(f'Товар ID={product.get("ID")} Name={product.get("PRODUCT_NAME")} не найден в каталоге товаров'
                         f' Битрикс24')
            return HttpResponse(status=415, content=f'Товар ID={product.get("ID")} Name={product.get("PRODUCT_NAME")} '
                                                    f'не найден в каталоге товаров Битрикс24')

        product_value: dict = dict()
        for key, value in product.items():
            if key in core_methods.codes_values:
                if core_methods.codes_values.get(key)[1] == 'int':
                    new_value = int(value)
                elif core_methods.codes_values.get(key)[1] == 'decimal':
                    new_value = round(decimal.Decimal(value), 2) or 0
                elif core_methods.codes_values.get(key)[1] == 'bool':
                    new_value = True if value == 'Y' else False
                else:
                    new_value = str(value)
                product_value[core_methods.codes_values.get(key)[0]] = new_value
        product_value['tax_sum'] = round(product_value.get('quantity') * product_value.get('price_exclusive') *
                                         product_value.get('tax') / 100, 2)
        product_value['sum'] = round(product_value.get('quantity') * product_value.get('price_account'), 2)
        product_value['portal'] = portal
        logger.debug(f'Свойства товарной позиции для обновления или создания:{NEW_STR}{product_value=}')

        prodtime, created = ProdTimeDeal.objects.update_or_create(product_id_b24=product_value.get('product_id_b24'),
                                                                  defaults=product_value)
        if created:
            prodtime.name_for_print = prodtime.name
        logger.info(f'Товарная позиция: {created=} {prodtime.id=}')

        logger.info(f'Работа с прямыми затратами, нормочасами, материалами, сроком производства из каталога')
        # Работа с прямыми затратами, нормочасами, материалами, сроком производства из каталога
        core_methods.set_fields_from_catalog_b24(prodtime, settings_portal, product_in_catalog)

        logger.info(f'Работа с прибылью')
        # Работа с прибылью
        logger.info(f'Свойство is_change_income имеет значение {prodtime.is_change_income}')
        if (deal_b24.properties.get(settings_portal.deal_field_code_income_res) and not prodtime.income
                and not prodtime.is_change_income):
            core_methods.calculation_income(prodtime, settings_portal)

        logger.info(f'Работа с эквивалентом')
        # Работа с эквивалентом
        logger.info(f'Свойство is_change_equivalent имеет значение {prodtime.is_change_equivalent}')
        if not prodtime.is_change_equivalent:
            equivalent_code = settings_portal.equivalent_code.lower().replace('_', '')
            if (equivalent_code not in product_in_catalog.properties
                    or not product_in_catalog.properties.get(equivalent_code)):
                prodtime.equivalent = 0
            else:
                equivalent_value = product_in_catalog.properties.get(equivalent_code).get('value')
                prodtime.equivalent = decimal.Decimal(equivalent_value)
            logger.info(f'Для свойства equivalent установлено значение {prodtime.equivalent}')
            prodtime.save()
        if prodtime.equivalent and prodtime.quantity:
            prodtime.equivalent_count = (prodtime.equivalent * prodtime.quantity)
        else:
            prodtime.equivalent_count = 0
        logger.info(f'Для свойства equivalent_count установлено значение {prodtime.equivalent_count}')

        prodtime.save()

    logger.info(f'Удаление лишних товарных позиций, которые удалены в сделке')
    # Удаление лишних товарных позиций, которые удалены в сделке
    products_in_db = ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id)
    for product in products_in_db:
        if not next((x for x in deal_b24.products if int(x.get('ID')) == product.product_id_b24), None):
            logger.info(f'{product.id} удален из БД, так как отсутствует в сделке')
            product.delete()

    logger.info(f'Успешное завершение обновление сделки')
    logger.info(f'{SEPARATOR}')
    return HttpResponse(status=200, content='Успешно')


@xframe_options_exempt
@csrf_exempt
def save(request):
    """Метод для сохранения изменений в таблице."""

    product_id = request.POST.get('id')
    type_field = request.POST.get('type')
    value = request.POST.get('value')

    if type_field == 'general-number':
        deal = get_object_or_404(Deal, pk=request.POST.get('id'))
        deal.general_number = value
        deal.save()
        return JsonResponse({"success": "Updated"})

    prodtime: ProdTimeDeal = get_object_or_404(ProdTimeDeal, pk=product_id)
    if type_field == 'name-for-print':
        prodtime.name_for_print = value
    if value:
        if type_field == 'count-days':
            prodtime.count_days = value
        if type_field == 'equivalent':
            prodtime.equivalent = value
            prodtime.equivalent_count = (decimal.Decimal(value) * prodtime.quantity)
            prodtime.is_change_equivalent = True
        if type_field == 'prodtime-str':
            prodtime.prodtime_str = value
            prodtime.is_change_prodtime_str = True
        if type_field == 'direct-costs':
            prodtime.direct_costs = value
            prodtime.is_change_direct_costs = True
        if type_field == 'direct-costs-fact':
            prodtime.direct_costs_fact = value
        if type_field == 'standard-hours':
            prodtime.standard_hours = value
            prodtime.is_change_standard_hours = True
        if type_field == 'standard-hours-fact':
            prodtime.standard_hours_fact = value
        if type_field == 'standard-hours2':
            prodtime.standard_hours2 = value
            prodtime.is_change_standard_hours2 = True
        if type_field == 'standard-hours2-fact':
            prodtime.standard_hours2_fact = value
        if type_field == 'standard-hours3':
            prodtime.standard_hours3 = value
            prodtime.is_change_standard_hours3 = True
        if type_field == 'standard-hours3-fact':
            prodtime.standard_hours3_fact = value
        if type_field == 'standard-hours4':
            prodtime.standard_hours4 = value
            prodtime.is_change_standard_hours4 = True
        if type_field == 'standard-hours4-fact':
            prodtime.standard_hours4_fact = value
        if type_field == 'materials':
            prodtime.materials = value
            prodtime.is_change_materials = True
        if type_field == 'materials-fact':
            prodtime.materials_fact = value
        if type_field == 'finish':
            if value == 'true':
                prodtime.finish = True
            if value == 'false':
                prodtime.finish = False
        if type_field == 'made':
            if value == 'true':
                prodtime.made = True
            if value == 'false':
                prodtime.made = False
    prodtime.save()

    return JsonResponse({"success": "Updated"})


@xframe_options_exempt
@csrf_exempt
def update_factory_number(request):
    """Метод для изменения заводского номера в элементе smart процесса."""
    product_id = request.POST.get('product_id')
    member_id = request.POST.get('member_id')
    value = request.POST.get('value')
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    try:
        prodtime = ProdTimeDeal.objects.get(pk=product_id)
    except ObjectDoesNotExist as ex:
        return JsonResponse({'result': 'error', 'info': ex.args[0]})

    try:
        smart_factory_number = SmartProcessB24(portal, settings_portal.smart_factory_number_id)
        fields = {'title': value, settings_portal.smart_factory_number_code: value}
        result = smart_factory_number.update_element(prodtime.smart_id_factory_number, fields)

        if 'item' not in result or result.get('item').get(settings_portal.smart_factory_number_code) != value:
            return JsonResponse({'result': 'error',
                                 'info': f'Не удалось обновить элемент с id = {prodtime.smart_id_factory_number} '
                                         f'smart процесса Заводские номера '})

        prodtime.factory_number = value
        prodtime.save()

    except RuntimeError as ex:
        return JsonResponse({'result': 'error', 'info': ex.args[0]})

    return JsonResponse({'result': 'success'})


@xframe_options_exempt
@csrf_exempt
def update_prodtime(request):
    """Метод для срока производства и подсчета максимального срока производства."""
    product_id = request.POST.get('product_id')
    member_id = request.POST.get('member_id')
    value = request.POST.get('value')
    deal_id = request.POST.get('deal_id')

    try:
        prodtime = ProdTimeDeal.objects.get(id=product_id)
        prodtime.prod_time = value if value else None
        prodtime.save()
        max_prodtime = core_methods.count_set_max_prodtime(member_id=member_id, deal_id=deal_id)

    except Exception as ex:
        return JsonResponse({'result': 'error', 'info': f'{ex.args[0]}: {ex.args[1]}'})

    return JsonResponse({'result': 'success', 'max_prodtime': max_prodtime})


@xframe_options_exempt
@csrf_exempt
def create(request):
    """Метод создания задачи и сделки."""

    products_id = request.POST.getlist('products_id[]')
    member_id = request.POST.get('member_id')
    deal_id = request.POST.get('deal_id')
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)
    general_msg = ''

    general_msg += f"Создание сделки:\n"
    result = core_methods.create_shipment_element(portal, settings_portal, deal_id, products_id)
    try:
        new_deal_id = int(result.get('info'))
    except ValueError:
        new_deal_id = 0
    general_msg += str(result.get('info'))
    general_msg += f"\nСоздание элемента смарт процесса:\n"
    result = core_methods.create_shipment_element(portal, settings_portal, deal_id, products_id, type_elem='smart')
    general_msg += str(result.get('info'))
    general_msg += f"\nСоздание задачи:\n"
    result = core_methods.create_shipment_task(portal, settings_portal, new_deal_id)
    general_msg += str(result.get('info'))
    core_methods.save_finish(products_id)

    return JsonResponse({'res': general_msg})


@xframe_options_exempt
@csrf_exempt
def update_direct_costs(request):
    """Метод обновления полей Прямые затраты, Нормочасы, Материалы и пересчет Прибыли."""
    name_fields = ['direct_costs', 'standard_hours', 'materials']

    member_id = request.POST.get('member_id')
    deal_id = int(request.POST.get('deal_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    products = ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id)
    if not products:
        return JsonResponse({
            'result': 'error',
            "info": 'В сделке отсутствуют товары.'
        })

    for product in products:
        productrow_in_deal = ProductRowB24(portal, product.product_id_b24)
        product_in_catalog = ProductInCatalogB24(
            portal, productrow_in_deal.id_in_catalog)

        for name_field in name_fields:
            code_field = getattr(settings_portal, name_field + '_code')
            value_field = product_in_catalog.properties.get(code_field)
            if type(value_field) is dict and 'value' in value_field:
                value = round(decimal.Decimal(value_field.get('value')), 2)
                setattr(product, name_field, value)
            else:
                setattr(product, name_field, 0)
            setattr(product, 'is_change_' + name_field, False)

        core_methods.calculation_income(product, settings_portal)

    return JsonResponse({
        'result': 'success',
        "info": 'Значения полей успешно обновлены из каталога товаров Битрикс24. Выполнен пересчет прибыли'
    })


@xframe_options_exempt
@csrf_exempt
def create_doc(request):
    """Метод создания документа по выбранному шаблону."""

    member_id = request.POST.get('member_id')
    template_id = request.POST.get('templ_id')
    deal_id = int(request.POST.get('deal_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    try:
        kp_number, next_kp_number = core_methods.update_kp_numbers_in_deal_b24(portal, settings_portal, deal_id,
                                                                               template_id)
    except RuntimeError as ex:
        return render(request, 'error.html', {'error_name': ex.args[0], 'error_description': ex.args[1]})

    products = ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id).values()
    products = sorted(products, key=lambda prod: prod.get('sort'))
    values = core_methods.fill_values_for_create_doc(settings_portal, template_id, products, kp_number)

    try:
        bx24_template_doc = TemplateDocB24(portal, 0)
        result = bx24_template_doc.create_docs(template_id, deal_id, values)
    except RuntimeError as ex:
        return render(request, 'error.html', {'error_name': ex.args[0], 'error_description': ex.args[1]})

    return JsonResponse({'result': result})


@xframe_options_exempt
@csrf_exempt
def copy_products(request):
    """Метод копирования товаров в каталог и сделку, создание артикулов."""

    member_id = request.POST.get('member_id')
    deal_id = request.POST.get('deal_id')
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal,
                                                        portal=portal)
    current_year = int(datetime.date.today().year)
    try:
        numeric = Numeric.objects.get(portal=portal, year=current_year)
        last_number_in_year = numeric.last_number
    except ObjectDoesNotExist:
        return JsonResponse({'result': 'error', 'info': f'Нумератор для года {current_year} отсутствует'})

    try:
        year_code = AssociativeYearNumber.objects.get(
            portal=portal, year=int(datetime.date.today().year)).year_code
    except ObjectDoesNotExist:
        return JsonResponse({'result': 'error', 'info': f'Код для года {datetime.date.today().year} отсутствует в '
                                                        f'таблице соответствия'})

    products = ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id)
    result_text = ''

    section_list = ListB24(portal, settings_portal.section_list_id)

    for product in products:
        try:
            product_row = ProductRowB24(portal, product.product_id_b24)
            product_in_catalog = ProductInCatalogB24(portal, product_row.id_in_catalog)
        except RuntimeError as ex:
            return JsonResponse({'result': 'error', 'info': str(ex.args[0]) + str(ex.args[1])})

        name = product.name_for_print
        section_id = product_in_catalog.properties.get('iblockSectionId')
        service = product_in_catalog.properties.get(settings_portal.service_code)

        try:
            filter_for_list = {settings_portal.real_section_code: section_id}
            elements_section_list = section_list.get_element_filter(filter_for_list)
        except RuntimeError:
            result_text += f'Для товара {name} Невозможно получить секции каталога для сопоставления\n'
            continue
        if not elements_section_list:
            new_section_id = settings_portal.default_section_id
        else:
            new_section_id = list(elements_section_list[0].get(settings_portal.copy_section_code).values())[0]

        is_auto_article = product_in_catalog.properties.get(settings_portal.is_auto_article_code)
        article = product_in_catalog.properties.get(settings_portal.article_code)
        section_number = product_in_catalog.properties.get(settings_portal.section_number_code)
        if service == 'Y':
            new_name = product.name_for_print
        else:
            if is_auto_article == 'N':
                result_text += f'Для товара {name} Присваивать артикул автоматически выключено\n'
                continue
            if not section_number or 'value' not in section_number:
                result_text += f'Для товара {name} Номер раздела не присвоен\n'
                continue
            section_number = product_in_catalog.properties.get(
                settings_portal.section_number_code).get('value')
            if article:
                result_text += f'Для товара {name} артикул уже существует\n'
                continue

            last_number_in_year += 1
            article = 'ПТ{}.{}{:06}{}'.format(section_number, year_code, last_number_in_year, '00')
            new_name = f"{product.name_for_print} ( {article} )"

        name_fields = ['direct_costs', 'standard_hours', 'materials', 'prodtime_str']
        for name_field in name_fields:
            code_field = getattr(settings_portal, name_field + '_code')
            if getattr(product, 'is_change_' + name_field):
                product_in_catalog.properties[code_field] = str(getattr(product, name_field))

        equivalent_code = ''.join(settings_portal.equivalent_code.split('_')).lower()
        if product.is_change_equivalent:
            product_in_catalog.properties[equivalent_code] = str(product.equivalent)

        product_in_catalog.properties['name'] = new_name
        product_in_catalog.properties['iblockSectionId'] = new_section_id
        product_in_catalog.properties['iblockSection'] = [new_section_id]
        product_in_catalog.properties['createdBy'] = settings_portal.responsible_id_copy_catalog
        product_in_catalog.properties['purchasingPrice'] = None
        product_in_catalog.properties['purchasingCurrency'] = 'RUB'
        factory_number_code = ''.join(settings_portal.factory_number_code.split('_')).lower()
        product_in_catalog.properties[factory_number_code] = {}
        product_in_catalog.properties[factory_number_code]['value'] = 'Y'
        product_in_catalog.properties[settings_portal.is_auto_article_code] = {}
        product_in_catalog.properties[settings_portal.service_code] = {}
        if service == 'Y':
            product_in_catalog.properties[settings_portal.is_auto_article_code]['valueEnum'] = 'Нет'
            product_in_catalog.properties[settings_portal.service_code]['value'] = 'Y'
        else:
            product_in_catalog.properties[settings_portal.is_auto_article_code]['valueEnum'] = 'Да'
            product_in_catalog.properties[settings_portal.service_code]['value'] = 'N'
            product_in_catalog.properties[settings_portal.article_code] = article
        del product_in_catalog.properties['id']
        product_in_catalog.check_and_update_properties()
        try:
            new_id_product_in_catalog = product_in_catalog.add().get('element').get('id')
            product_row.properties['productId'] = new_id_product_in_catalog
            product_row.properties['productName'] = new_name
            del product_row.properties['id']
            product_row.update(product.product_id_b24)
        except RuntimeError as ex:
            return JsonResponse({'result': 'error', 'info': f'{ex.args[0]} {ex.args[1]}'})
        if service == 'Y':
            result_text += f'Товар {name} скопирован как услуга\n'
        else:
            result_text += f'Для товара {name} артикул присвоен {article}\n'

        prodtime = ProdTimeDeal.objects.get(portal=portal, product_id_b24=product_row.id)
        prodtime.name_for_print = new_name
        prodtime.save()

    numeric.last_number = last_number_in_year
    numeric.save()
    return JsonResponse({'result': 'success', 'info': result_text})


@xframe_options_exempt
@csrf_exempt
def write_factory_number(request):
    """Метод создания заводских номеров."""

    member_id = request.POST.get('member_id')
    deal_id = request.POST.get('deal_id')
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    try:
        deal = Deal.objects.get(deal_id=deal_id, portal=portal)
        products = list(ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id))
        products = sorted(products, key=lambda prod: prod.sort)
        deal_b24 = DealB24(portal, deal_id)
    except Exception as ex:
        return JsonResponse({'result': 'error', 'info': ex.args[0]})

    i = 0
    result_work = ''
    for product in products:
        try:
            if product.quantity == 1:
                continue
            productrow = ProductRowB24(portal, product.product_id_b24)
            product_in_catalog = ProductInCatalogB24(portal, productrow.id_in_catalog)
            product_factory_number_value = product_in_catalog.properties.get(settings_portal.factory_number_code)
            if not product_factory_number_value or product_factory_number_value.get('valueEnum') == 'Нет':
                continue
            if product.factory_number and product.smart_id_factory_number:
                continue
            if product.quantity > 1 and (int(product.quantity) == product.quantity):
                start_sort = product.sort
                new_productrow = ProductRowB24(portal, 0)
                for count in range(int(product.quantity)):
                    fields = {
                        'ownerId': deal_id,
                        'ownerType': 'D',
                        'productId': product_in_catalog.id,
                        'price': str(product.price),
                        'quantity': 1,
                        'discountTypeId': product.bonus_type_id,
                        'discountRate': str(product.bonus),
                        'discountSum': str(product.bonus_sum / product.quantity),
                        'taxRate': str(product.tax),
                        'measureCode': product.measure_code,
                        'measureName': product.measure_name,
                        'sort': start_sort
                    }
                    result = new_productrow.add(fields)
                    if product.income:
                        income = round(product.income / product.quantity, 2)
                    else:
                        income = 0
                    new_product = ProdTimeDeal.objects.create(
                        product_id_b24=result.get('productRow').get('id'),
                        name=product.name,
                        name_for_print=product.name_for_print,
                        price=product.price,
                        price_account=product.price_account,
                        price_exclusive=product.price_exclusive,
                        price_netto=product.price_netto,
                        price_brutto=product.price_brutto,
                        quantity=1,
                        income=income,
                        is_change_income=product.is_change_income,
                        measure_code=product.measure_code,
                        measure_name=product.measure_name,
                        bonus_type_id=product.bonus_type_id,
                        bonus=product.bonus,
                        bonus_sum=round(decimal.Decimal(result.get(
                            'productRow').get('discountSum')), 2),
                        tax=product.tax,
                        tax_included=product.tax_included,
                        tax_sum=product.price_exclusive * product.tax / 100,
                        sum=product.price_account,
                        sort=start_sort,
                        prod_time=product.prod_time,
                        count_days=product.count_days,
                        equivalent=product.equivalent,
                        equivalent_count=product.equivalent,
                        is_change_equivalent=product.is_change_equivalent,
                        finish=product.finish,
                        made=product.made,
                        deal_id=deal_id,
                        direct_costs=product.direct_costs,
                        direct_costs_fact=product.direct_costs_fact,
                        is_change_direct_costs=product.is_change_direct_costs,
                        materials=product.materials,
                        materials_fact=product.materials_fact,
                        is_change_materials=product.is_change_materials,
                        standard_hours=product.standard_hours,
                        standard_hours_fact=product.standard_hours_fact,
                        is_change_standard_hours=product.is_change_standard_hours,
                        standard_hours2=product.standard_hours2,
                        standard_hours2_fact=product.standard_hours2_fact,
                        is_change_standard_hours2=product.is_change_standard_hours2,
                        standard_hours3=product.standard_hours3,
                        standard_hours3_fact=product.standard_hours3_fact,
                        is_change_standard_hours3=product.is_change_standard_hours3,
                        standard_hours4=product.standard_hours4,
                        standard_hours4_fact=product.standard_hours4_fact,
                        is_change_standard_hours4=product.is_change_standard_hours4,
                        prodtime_str=product.prodtime_str,
                        is_change_prodtime_str=product.is_change_prodtime_str,
                        portal=portal,
                    )
                    products.append(new_product)
                productrow.delete()
                product.delete()
                continue
        except Exception as ex:
            return JsonResponse({
                'result': 'error',
                'info': f'Ошибка: {ex.args[0]}. Описание ошибки: {ex.args[1]} для товара {product.name = }'
            })

    products = list(ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id))
    products = sorted(products, key=lambda prod: prod.sort)
    for product in products:
        try:
            productrow = ProductRowB24(portal, product.product_id_b24)
            product_in_catalog = ProductInCatalogB24(portal, productrow.id_in_catalog)
            product_factory_number_value = product_in_catalog.properties.get(settings_portal.factory_number_code)
            if not product_factory_number_value or product_factory_number_value.get('valueEnum') == 'Нет':
                continue
            if product.factory_number and product.smart_id_factory_number:
                continue
            product.factory_number = '{}{}'.format(
                deal.general_number,
                str(deal.last_factory_number + 1)
            )
            product.save()
            deal.last_factory_number += 1
            deal.save()
            i += 1
            smart_factory_number = SmartProcessB24(portal, settings_portal.smart_factory_number_id)
            fields = {
                'title': product.factory_number,
                'assigned_by_id': deal_b24.properties.get('ASSIGNED_BY_ID'),
                'parentId2': deal_id,
                'company_id': deal_b24.company_id,
                settings_portal.smart_factory_number_code: product.factory_number,
            }
            result = smart_factory_number.create_element(fields)
            if 'item' not in result:
                continue
            product.smart_id_factory_number = int(result.get('item').get('id'))
            product.save()

            fields = {
                'ownerId': product.smart_id_factory_number,
                'ownerType': settings_portal.smart_factory_number_short_code,
                'productId': product_in_catalog.id,
                'quantity': 1,
            }
            productrow.add(fields)
            result_work += f'\nДля товара {product.name} установлен заводской номер {product.factory_number}'
        except Exception as ex:
            return JsonResponse({'result': 'error', 'info': f'{ex.args[0]} для товара {product.name = }'})

    if i == 0:
        return JsonResponse({'result': 'success', 'info': 'Нет товаров, в которых можно установить заводские номера'})
    else:
        return JsonResponse({'result': 'success', 'info': result_work})


@xframe_options_exempt
@csrf_exempt
def send_equivalent(request):
    """Метод отправки суммарного эквивалента в сделку."""

    member_id = request.POST.get('member_id')
    deal_id = int(request.POST.get('deal_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    sum_equivalent = core_methods.count_sum_equivalent(ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id))
    if sum_equivalent:
        sum_equivalent = float(sum_equivalent)
        deal = DealB24(portal, deal_id)
        fields = {settings_portal.sum_equivalent_code: sum_equivalent}
        deal.update(fields)

    return JsonResponse({'result': sum_equivalent})


@xframe_options_exempt
@csrf_exempt
def export_excel(request):
    """Метод экспорта в excel."""

    member_id = request.GET.get('member_id')
    deal_id = int(request.GET.get('deal_id'))
    portal: Portals = create_portal(member_id)

    products = ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id)
    products = sorted(products, key=lambda prod: prod.sort)

    deal = DealB24(portal, deal_id)

    try:
        company = CompanyB24(portal, deal.company_id)
        company_name = company.name
    except RuntimeError:
        company_name = 'Не удалось получить наименование компании'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
        deal.properties.get('UF_CRM_1661507258282') or 'report')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Prodtime'

    columns = [
        'Товар',
        'Кол-во',
        'Кол-во раб. дней',
        'Срок производства',
        'Заводской номер',
        'Нормочасы План за 1 ед',
        'Нормочасы Факт за 1 ед',
        'Материалы План за 1 ед без НДС',
        'Материалы Факт за 1 ед без НДС',
        'Прямые затраты План за 1 ед без НДС',
        'Прямые затраты Факт за 1 ед без НДС',
        'Срок производства из каталога',
    ]

    worksheet.merge_cells('B1:D1')
    worksheet.merge_cells('B2:D2')
    worksheet.cell(1, 1).value = 'ИД сделки:'
    worksheet.cell(1, 2).value = deal_id
    worksheet.cell(1, 2).alignment = Alignment(horizontal='left')
    worksheet.cell(2, 1).value = 'Компания сделки:'
    worksheet.cell(2, 2).value = company_name

    row_num = 4
    worksheet.column_dimensions['A'].width = 60
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 17
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 10
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 10
    worksheet.column_dimensions['H'].width = 10
    worksheet.column_dimensions['I'].width = 10
    worksheet.column_dimensions['J'].width = 10
    worksheet.column_dimensions['K'].width = 10
    worksheet.column_dimensions['L'].width = 15

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for product in products:
        row_num += 1
        prod_time = (product.prod_time.strftime('%d.%m.%Y') if product.prod_time else '')

        row = [
            product.name_for_print,
            product.quantity,
            product.count_days,
            prod_time,
            product.factory_number,
            product.standard_hours,
            product.standard_hours_fact,
            product.materials,
            product.materials_fact,
            product.direct_costs,
            product.direct_costs_fact,
            product.prodtime_str,
        ]

        for col_num, cell_value in enumerate(row, 1):
            worksheet.row_dimensions[row_num].height = 30
            cell = worksheet.cell(row=row_num, column=col_num)
            alignment = Alignment(vertical='center', wrap_text=True)
            cell.alignment = alignment
            cell.value = cell_value

    workbook.save(response)

    return response


class ProdTimeDealViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ProdTimeDeal.objects.all()
    serializer_class = ProdTimeDealSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = {
        'deal_id': ['exact'],
        'updated': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'portal': ['exact'],
    }


class ObjB24Old:
    """Класс объекта Битрикс24"""

    def __init__(self, portal: Portals):
        self.portal = portal
        self.bx24 = Bitrix24(portal.name)
        self.bx24._access_token = portal.auth_id

    @staticmethod
    def _check_error(result):
        if 'error' in result:
            raise RuntimeError(result['error'], result['error_description'])
        elif 'result' in result:
            return result['result']
        else:
            raise RuntimeError('Error', 'No description error')


class ProductB24Old(ObjB24Old):
    """Класс товара каталога Битрикс24."""

    def __init__(self, portal, product_id):
        super(ProductB24Old, self).__init__(portal)
        self.id = product_id
        self.productrow = None
        self.id_catalog = self.get_product_catalog_id()
        self.props = self.get_properties()

    def get_product_catalog_id(self):
        method_rest = 'crm.item.productrow.get'
        params = {'id': self.id}
        result = self.bx24.call(method_rest, params)
        self.productrow = self._check_error(result).get('productRow')
        return self.productrow.get('productId')

    def get_properties(self):
        """Метод получения свойств товара каталога."""
        method_rest = 'crm.product.get'
        params = {'id': self.id_catalog}
        result = self.bx24.call(method_rest, params)
        return self._check_error(result)

    def add_catalog(self):
        """Метод добавления товара в каталог."""
        method_rest = 'crm.product.add'
        params = {'fields': self.props}
        result = self.bx24.call(method_rest, params)
        return self._check_error(result)

    def update(self, product_id):
        """Метод изменения товарной позиции."""
        method_rest = 'crm.item.productrow.update'
        params = {
            'id': product_id,
            'fields': self.productrow
        }
        result = self.bx24.call(method_rest, params)
        return self._check_error(result)
