import datetime
import decimal

import core.methods as core_methods

from typing import Dict, Any

from django.db.models import Sum
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist, BadRequest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from core.methods import initial_check, get_current_user, calculation_income
from core.models import Portals
from dealcard.models import ProdTimeDeal
from settings.models import SettingsPortal, Numeric, AssociativeYearNumber
from .models import ProdTimeQuote

from core.bitrix24.bitrix24 import (
    QuoteB24, TemplateDocB24, ProductB24, CompanyB24, DealB24, ProductRowB24,
    ListB24, ProductInCatalogB24, create_portal)


@xframe_options_exempt
@csrf_exempt
def index(request):
    """Метод страницы Карточка предложения."""
    template: str = 'quotecard/index.html'
    title: str = 'Страница карточки сделки'

    try:
        member_id, quote_id, auth_id = initial_check(request, 'quote_id')
    except BadRequest:
        return render(request, 'error.html', {
            'error_name': 'QueryError',
            'error_description': 'Неизвестный тип запроса'
        })
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal,
                                                        portal=portal)
    user_info = get_current_user(request, auth_id, portal,
                                 settings_portal.is_admin_code)
    year = int(datetime.date.today().year)
    Numeric.objects.get_or_create(portal=portal, year=year,
                                  defaults={'last_number': 0})

    try:
        quote = QuoteB24(portal, quote_id)
        quote.get_all_products()
        template_doc = TemplateDocB24(portal, 0)
        template_doc.get_all_templates(parent='quote')
        templates = ([{'id': templ['id'], 'name': templ['name']}
                      for templ in template_doc.templates.values()])
    except RuntimeError as ex:
        return render(request, 'error.html', {
            'error_name': ex.args[0],
            'error_description': ex.args[1]
        })

    if not quote.deal_id or int(quote.deal_id) == 0:
        return render(request, 'error.html', {
            'error_name': 'Отсутствует привязка к сделке',
            'error_description': 'Привяжите сделку и обновите страницу.'
        })

    products = []
    for product in quote.products:
        product_value: Dict[str, Any] = {
            'product_id_b24': int(product['ID']),
            'name': product['PRODUCT_NAME'],
            'name_for_print': product['PRODUCT_NAME'],
            'price': round(decimal.Decimal(product['PRICE']), 2),
            'price_acc': round(decimal.Decimal(product['PRICE_ACCOUNT']), 2),
            'price_exs': round(decimal.Decimal(product['PRICE_EXCLUSIVE']), 2),
            'price_netto': round(decimal.Decimal(product['PRICE_NETTO']), 2),
            'price_brutto': round(decimal.Decimal(product['PRICE_BRUTTO']), 2),
            'quantity': round(decimal.Decimal(product['QUANTITY']), 2),
            'measure_code': int(product['MEASURE_CODE']),
            'measure_name': product['MEASURE_NAME'],
            'bonus_type_id': int(product['DISCOUNT_TYPE_ID']),
            'bonus': round(decimal.Decimal(product['DISCOUNT_RATE']), 2),
            'bonus_sum': round(decimal.Decimal(product['DISCOUNT_SUM']), 2),
            'tax': 0 if not product['TAX_RATE'] else round(
                decimal.Decimal(product['TAX_RATE']), 2),
            'tax_included': True if product['TAX_INCLUDED'] == 'Y' else False,
            'sort': int(product['SORT']),
        }
        product_value['tax_sum'] = round(product_value['quantity'] *
                                         product_value['price_exs'] *
                                         product_value['tax'] / 100, 2)
        product_value['sum'] = round(product_value['quantity']
                                     * product_value['price_acc'], 2)
        try:
            prodtime: ProdTimeQuote = ProdTimeQuote.objects.get(
                product_id_b24=product_value['product_id_b24']
            )
            product_value['name_for_print'] = prodtime.name_for_print
            product_value['prod_time'] = prodtime.prod_time
            product_value['count_days'] = str(prodtime.count_days).replace(',',
                                                                           '.')
            product_value['finish'] = prodtime.finish
            product_value['made'] = prodtime.made
            product_value['smart_id_factory_number'] = (
                prodtime.smart_id_factory_number)
            product_value['factory_number'] = prodtime.factory_number
            product_value['id'] = prodtime.pk
            prodtime.name = product_value['name']
            prodtime.price = product_value['price']
            prodtime.price_account = product_value['price_acc']
            prodtime.price_exclusive = product_value['price_exs']
            prodtime.price_netto = product_value['price_netto']
            prodtime.price_brutto = product_value['price_brutto']
            prodtime.quantity = product_value['quantity']
            prodtime.measure_code = product_value['measure_code']
            prodtime.measure_name = product_value['measure_name']
            prodtime.bonus_type_id = product_value['bonus_type_id']
            prodtime.bonus = product_value['bonus']
            prodtime.bonus_sum = product_value['bonus_sum']
            prodtime.tax = product_value['tax']
            prodtime.tax_included = product_value['tax_included']
            prodtime.tax_sum = product_value['tax_sum']
            prodtime.sum = product_value['sum']
            prodtime.sort = product_value['sort']
            prodtime.save()
        except ObjectDoesNotExist:
            prodtime: ProdTimeQuote = ProdTimeQuote.objects.create(
                product_id_b24=product_value['product_id_b24'],
                name=product_value['name'],
                name_for_print=product_value['name_for_print'],
                price=product_value['price'],
                price_account=product_value['price_acc'],
                price_exclusive=product_value['price_exs'],
                price_netto=product_value['price_netto'],
                price_brutto=product_value['price_brutto'],
                quantity=product_value['quantity'],
                measure_code=product_value['measure_code'],
                measure_name=product_value['measure_name'],
                bonus_type_id=product_value['bonus_type_id'],
                bonus=product_value['bonus'],
                bonus_sum=product_value['bonus_sum'],
                tax=product_value['tax'],
                tax_included=product_value['tax_included'],
                tax_sum=product_value['tax_sum'],
                sum=product_value['sum'],
                sort=product_value['sort'],
                quote_id=quote_id,
                portal=portal,
            )
            product_value['id'] = prodtime.pk

        # Проверка товара на каталог
        try:
            if product["PRODUCT_ID"] == 0:
                raise RuntimeError('Product not catalog',
                                   'Product not catalog')
            product_in_catalog = ProductB24(portal, product["PRODUCT_ID"])
        except RuntimeError as ex:
            return render(request, 'error.html', {
                'error_name': ex.args[0],
                'error_description': f'{ex.args[1]} для товара '
                                     f'{prodtime.name}'
            })

        # Работа с эквивалентом
        if prodtime.is_change_equivalent:
            product_value['equivalent'] = str(prodtime.equivalent).replace(',', '.')
        else:
            equivalent_code = settings_portal.equivalent_code
            if (not product_in_catalog.properties or equivalent_code not in product_in_catalog.properties
                    or not product_in_catalog.properties.get(equivalent_code)):
                product_value['equivalent'] = ''
                prodtime.equivalent = 0
            else:
                product_value['equivalent'] = list(product_in_catalog.properties.get(equivalent_code).values())[1]
                prodtime.equivalent = decimal.Decimal(product_value['equivalent'])
            prodtime.save()
        if prodtime.equivalent and prodtime.quantity:
            prodtime.equivalent_count = (prodtime.equivalent * prodtime.quantity)
            prodtime.save()

        # Работа с прибылью
        if not prodtime.income and not prodtime.is_change_income:
            calculation_income(prodtime, settings_portal)

        # income_code = settings_portal.income_code.upper().replace('Y', 'Y_')
        # if income_code not in product_in_catalog.properties or not product_in_catalog.properties.get(income_code):
        #     product_value['income'] = ''
        #     prodtime.income = 0
        # else:
        #     income_value_per_field = product_in_catalog.properties.get(income_code)
        #     income_value_per = round(decimal.Decimal(income_value_per_field.get('value')), 2)
        #     income_result = round(prodtime.sum * income_value_per / 100, 2)
        #     prodtime.income = income_result
        prodtime.save()

        # Работа со сроком производства
        if prodtime.is_change_prodtime_str:
            product_value['prodtime_str'] = prodtime.prodtime_str
        else:
            try:
                if product["PRODUCT_ID"] == 0:
                    raise RuntimeError('Product not catalog', 'Product not catalog')
                product_in_catalog = ProductInCatalogB24(portal, product["PRODUCT_ID"])
            except RuntimeError as ex:
                return render(request, 'error.html', {
                    'error_name': ex.args[0],
                    'error_description': f'{ex.args[1]} для товара {prodtime.name}'
                })
            prodtime_str_code = settings_portal.prodtime_str_code
            if (not product_in_catalog.properties
                    or prodtime_str_code not in product_in_catalog.properties
                    or not product_in_catalog.properties.get(
                        prodtime_str_code)):
                product_value['prodtime_str'] = ''
                prodtime.prodtime_str = ''
            else:
                product_value['prodtime_str'] = (
                    product_in_catalog.properties.get(prodtime_str_code).
                    get('value'))
                prodtime.prodtime_str = product_value['prodtime_str']
            prodtime.save()
        products.append(product_value)

    products_in_db = ProdTimeQuote.objects.filter(portal=portal,
                                                  quote_id=quote_id)
    for product in products_in_db:
        if not next((x for x in products if
                     x['product_id_b24'] == product.product_id_b24), None):
            product.delete()

    # sum_equivalent = ProdTimeQuote.objects.filter(
    #     portal=portal, quote_id=quote_id).aggregate(Sum('equivalent_count'))
    sum_equivalent = core_methods.count_sum_equivalent(ProdTimeQuote.objects.filter(portal=portal, quote_id=quote_id))

    products = sorted(products, key=lambda prod: prod.get('sort'))

    context = {
        'title': title,
        'products': products,
        'sum_equivalent': sum_equivalent,
        'member_id': member_id,
        'quote_id': quote_id,
        'deal_id': int(quote.deal_id),
        'templates': templates,
        'user': user_info
    }
    response = render(request, template, context)
    if auth_id:
        response.set_cookie(key='user_id', value=user_info.get('user_id'))
    return render(request, template, context)


@xframe_options_exempt
@csrf_exempt
def save(request):
    """Метод для сохранения изменений в таблице."""
    product_id = request.POST.get('id')
    type_field = request.POST.get('type')
    value = request.POST.get('value')

    prodtime: ProdTimeQuote = get_object_or_404(ProdTimeQuote, pk=product_id)
    if type_field == 'name-for-print':
        prodtime.name_for_print = value
    if type_field == 'prod-time':
        prodtime.prod_time = value if value else None
    if type_field == 'count-days' and value:
        prodtime.count_days = value
    if type_field == 'prodtime-str':
        if value:
            prodtime.prodtime_str = value
            prodtime.is_change_prodtime_str = True
    if type_field == 'equivalent':
        if value:
            prodtime.equivalent = value
            prodtime.equivalent_count = (decimal.Decimal(value) * prodtime.quantity)
            prodtime.is_change_equivalent = True
    prodtime.save()

    return JsonResponse({"success": "Updated"})


@xframe_options_exempt
@csrf_exempt
def create_doc(request):
    """Метод создания документа по выбранному шаблону."""

    member_id = request.POST.get('member_id')
    template_id = request.POST.get('templ_id')
    quote_id = int(request.POST.get('quote_id'))
    deal_id = int(request.POST.get('deal_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    try:
        kp_number, next_kp_number = core_methods.update_kp_numbers_in_deal_b24(portal, settings_portal, deal_id,
                                                                               template_id)
    except RuntimeError as ex:
        return render(request, 'error.html', {'error_name': ex.args[0], 'error_description': ex.args[1]})

    products = ProdTimeQuote.objects.filter(portal=portal, quote_id=quote_id).values()
    products = sorted(products, key=lambda prod: prod.get('sort'))
    values = core_methods.fill_values_for_create_doc(settings_portal, template_id, products, kp_number)

    try:
        bx24_template_doc = TemplateDocB24(portal, 0)
        result = bx24_template_doc.create_docs(template_id, quote_id, values, parent='quote')
        if next_kp_number:
            quote = QuoteB24(portal, quote_id)
            fields = {'UF_CRM_62D8007A7602C': next_kp_number}
            quote.update(fields)
    except RuntimeError as ex:
        return render(request, 'error.html', {'error_name': ex.args[0], 'error_description': ex.args[1]})

    return JsonResponse({'result': result})


@xframe_options_exempt
@csrf_exempt
def create_articles(request):
    """Метод создания артикулов."""

    member_id = request.POST.get('member_id')
    quote_id = int(request.POST.get('quote_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal,
                                                        portal=portal)
    current_year = int(datetime.date.today().year)
    try:
        numeric = Numeric.objects.get(portal=portal, year=current_year)
        last_number_in_year = numeric.last_number
    except ObjectDoesNotExist:
        return JsonResponse({'result': 'error', 'info': f'Нумератор для года '
                                                        f'{current_year} '
                                                        f'отсутствует'})

    try:
        year_code = AssociativeYearNumber.objects.get(
            portal=portal, year=int(datetime.date.today().year)).year_code
    except ObjectDoesNotExist:
        return JsonResponse({'result': 'error',
                             'info': f'Код для года '
                                     f'{datetime.date.today().year} '
                                     f'отсутствует в таблице соответствия'})

    try:
        quote = QuoteB24(portal, quote_id)
        last_kp_number = int(quote.properties.get('UF_CRM_62D8007A7602C'))
    except RuntimeError as ex:
        return JsonResponse({'result': 'error',
                             'info': str(ex.args[0]) + str(ex.args[1])})

    products = ProdTimeQuote.objects.filter(portal=portal, quote_id=quote_id)
    result_text = ''

    section_list = ListB24(portal, settings_portal.section_list_id)

    for product in products:
        try:
            product_row = ProductRowB24(portal, product.product_id_b24)
            product_in_catalog = ProductInCatalogB24(
                portal, product_row.id_in_catalog)
        except RuntimeError as ex:
            return JsonResponse({'result': 'error',
                                 'info': str(ex.args[0]) + str(ex.args[1])})

        name = product.name_for_print
        section_id = product_in_catalog.properties.get('iblockSectionId')
        service = product_in_catalog.properties.get(
            settings_portal.service_code)

        try:
            filter_for_list = {settings_portal.real_section_code: section_id}
            elements_section_list = section_list.get_element_filter(
                filter_for_list)
        except RuntimeError:
            result_text += (f'Для товара {name} Невозможно получить секции '
                            f'каталога для сопоставления\n')
            continue
        if not elements_section_list:
            new_section_id = settings_portal.default_section_id
        else:
            new_section_id = list(elements_section_list[0].get(
                settings_portal.copy_section_code).values())[0]

        is_auto_article = product_in_catalog.properties.get(
            settings_portal.is_auto_article_code)
        article = product_in_catalog.properties.get(
            settings_portal.article_code)
        section_number = product_in_catalog.properties.get(
            settings_portal.section_number_code)
        if service == 'Y':
            new_name = product.name_for_print
        else:
            if is_auto_article == 'N':
                result_text += (f'Для товара {name} Присваивать артикул '
                                f'автоматически выключено\n')
                continue
            if not section_number or 'value' not in section_number:
                result_text += (f'Для товара {name} Номер раздела '
                                f'не присвоен\n')
                continue
            section_number = product_in_catalog.properties.get(
                settings_portal.section_number_code).get('value')
            if article:
                result_text += f'Для товара {name} артикул уже существует\n'
                continue

            last_number_in_year += 1
            kp_number = last_kp_number + 1
            article = 'ПТ{}.{}{:06}{:02}'.format(section_number, year_code,
                                                 last_number_in_year,
                                                 kp_number)
            new_name = f"{product.name_for_print} ( {article} )"

        equivalent_code = ''.join(settings_portal.equivalent_code.split('_')).lower()
        if product.is_change_equivalent:
            product_in_catalog.properties[equivalent_code] = str(product.equivalent)

        product_in_catalog.properties['name'] = new_name
        product_in_catalog.properties['iblockSectionId'] = new_section_id
        product_in_catalog.properties['iblockSection'] = [new_section_id]
        product_in_catalog.properties['createdBy'] = (
            settings_portal.responsible_id_copy_catalog)
        product_in_catalog.properties['purchasingPrice'] = None
        product_in_catalog.properties['purchasingCurrency'] = 'RUB'
        factory_number_code = ''.join(
            settings_portal.factory_number_code.split('_')).lower()
        product_in_catalog.properties[factory_number_code] = {}
        product_in_catalog.properties[factory_number_code]['value'] = 'Y'
        product_in_catalog.properties[
            settings_portal.is_auto_article_code] = {}
        product_in_catalog.properties[
            settings_portal.service_code] = {}
        if service == 'Y':
            product_in_catalog.properties[
                settings_portal.is_auto_article_code]['valueEnum'] = 'Нет'
            product_in_catalog.properties[
                settings_portal.service_code]['value'] = 'Y'
        else:
            product_in_catalog.properties[
                settings_portal.is_auto_article_code]['valueEnum'] = 'Да'
            product_in_catalog.properties[
                settings_portal.service_code]['value'] = 'N'
            product_in_catalog.properties[
                settings_portal.article_code] = article
        del product_in_catalog.properties['id']

        product_in_catalog.check_and_update_properties()

        try:
            new_id_product_in_catalog = product_in_catalog.add().get(
                'element').get('id')
            product_row.properties['productId'] = new_id_product_in_catalog
            product_row.properties['productName'] = new_name
            del product_row.properties['id']
            product_row.update(product.product_id_b24)
        except RuntimeError as ex:
            return JsonResponse({'result': 'error',
                                 'info': f'{ex.args[0]} {ex.args[1]}'})
        if service == 'Y':
            result_text += f'Товар {name} скопирован как услуга\n'
        else:
            result_text += f'Для товара {name} артикул присвоен {article}\n'

        prodtime = ProdTimeQuote.objects.get(portal=portal,
                                             product_id_b24=product_row.id)
        prodtime.name_for_print = new_name
        prodtime.save()

    numeric.last_number = last_number_in_year
    numeric.save()

    return JsonResponse({'result': 'success', 'info': result_text})


@xframe_options_exempt
@csrf_exempt
def copy_products(request):
    """Метод копирования товаров в каталог и предложение."""

    member_id = request.POST.get('member_id')
    quote_id = int(request.POST.get('quote_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal,
                                                        portal=portal)

    result_copy = 'error'
    products = ProdTimeQuote.objects.filter(portal=portal, quote_id=quote_id)

    for product in products:
        if product.name == product.name_for_print:
            continue
        try:
            productrow = ProductRowB24(portal, product.product_id_b24)
            product_in_catalog = ProductB24(portal, productrow.id_in_catalog)
            section_id = product_in_catalog.properties.get('SECTION_ID')
            filter_for_list = {settings_portal.real_section_code: section_id}
            element_list = ListB24(portal, settings_portal.section_list_id).get_element_filter(filter_for_list)
            if not element_list:
                new_section_id = settings_portal.default_section_id
            else:
                new_section_id = list(element_list[0].get(settings_portal.copy_section_code).values())[0]
            if product.is_change_equivalent:
                product_in_catalog.properties[settings_portal.equivalent_code] = {}
                product_in_catalog.properties[settings_portal.equivalent_code]['value'] = str(product.equivalent)
            if product.is_change_prodtime_str:
                product_in_catalog.properties[settings_portal.prodtime_str_code] = {}
                product_in_catalog.properties[settings_portal.prodtime_str_code]['value'] = str(product.prodtime_str)
            product_in_catalog.properties['NAME'] = product.name_for_print
            product_in_catalog.properties['SECTION_ID'] = new_section_id
            product_in_catalog.properties['CREATED_BY'] = (settings_portal.responsible_id_copy_catalog)
            product_in_catalog.properties['PRICE'] = None
            product_in_catalog.properties[settings_portal.price_with_tax_code] = None
            del product_in_catalog.properties['ID']
            new_id_product_in_catalog = product_in_catalog.add_catalog()
            productrow.properties['productId'] = new_id_product_in_catalog
            productrow.properties['productName'] = product.name_for_print
            del productrow.properties['id']
            result = productrow.update(product.product_id_b24)
            if 'productRow' in result:
                result_copy = 'ok'

        except RuntimeError as ex:
            return render(request, 'error.html', {
                'error_name': ex.args[0],
                'error_description': ex.args[1]
            })

    return JsonResponse({'result': result_copy})


@xframe_options_exempt
@csrf_exempt
def send_equivalent(request):
    """Метод отправки суммарного эквивалента в предложение."""

    member_id = request.POST.get('member_id')
    quote_id = int(request.POST.get('quote_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    sum_equivalent = core_methods.count_sum_equivalent(ProdTimeQuote.objects.filter(portal=portal, quote_id=quote_id))
    if sum_equivalent:
        sum_equivalent = float(sum_equivalent)
        quote = QuoteB24(portal, quote_id)
        result = quote.send_equivalent(settings_portal.sum_equivalent_quote_code, sum_equivalent)
        print(result)

    return JsonResponse({'result': sum_equivalent})


@xframe_options_exempt
@csrf_exempt
def send_products(request):
    """Метод передачи товаров из предложения в связанную сделку."""
    member_id = request.POST.get('member_id')
    deal_id = int(request.POST.get('deal_id'))
    quote_id = int(request.POST.get('quote_id'))
    portal: Portals = create_portal(member_id)
    settings_portal: SettingsPortal = get_object_or_404(SettingsPortal, portal=portal)

    try:
        quote = QuoteB24(portal, quote_id)
        quote.get_all_products()
        for product in quote.products:
            keys_for_del = ['ID', 'OWNER_ID', 'OWNER_TYPE']
            for key in keys_for_del:
                del product[key]
        deal = DealB24(portal, deal_id)
        deal.set_products(quote.products)
        deal.get_all_products()
        quote.get_all_products()
        ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id).delete()
        for count, product in enumerate(quote.products):
            prodtime_in_quote = ProdTimeQuote.objects.get(
                portal=portal, quote_id=quote_id,
                product_id_b24=int(product.get('ID')))
            ProdTimeDeal.objects.create(
                product_id_b24=deal.products[count].get('ID'),
                name=prodtime_in_quote.name,
                name_for_print=prodtime_in_quote.name_for_print,
                price=prodtime_in_quote.price,
                price_account=prodtime_in_quote.price_account,
                price_exclusive=prodtime_in_quote.price_exclusive,
                price_netto=prodtime_in_quote.price_netto,
                price_brutto=prodtime_in_quote.price_brutto,
                quantity=prodtime_in_quote.quantity,
                measure_code=prodtime_in_quote.measure_code,
                measure_name=prodtime_in_quote.measure_name,
                bonus_type_id=prodtime_in_quote.bonus_type_id,
                bonus=prodtime_in_quote.bonus,
                bonus_sum=prodtime_in_quote.bonus_sum,
                tax=prodtime_in_quote.tax,
                tax_included=prodtime_in_quote.tax_included,
                tax_sum=prodtime_in_quote.tax_sum,
                sum=prodtime_in_quote.sum,
                sort=prodtime_in_quote.sort,
                deal_id=deal.id,
                equivalent=prodtime_in_quote.equivalent,
                equivalent_count=prodtime_in_quote.equivalent_count,
                income=0,
                is_change_income=False,
                is_change_equivalent=prodtime_in_quote.is_change_equivalent,
                prod_time=prodtime_in_quote.prod_time,
                count_days=prodtime_in_quote.count_days,
                prodtime_str=prodtime_in_quote.prodtime_str,
                is_change_prodtime_str=(
                    prodtime_in_quote.is_change_prodtime_str),
                portal=portal,
            )
    except RuntimeError as ex:
        return JsonResponse({'result': 'error', 'error_name': ex.args[0], 'error_description': ex.args[1]})

    sum_equivalent = core_methods.count_sum_equivalent(ProdTimeDeal.objects.filter(portal=portal, deal_id=deal_id))
    if sum_equivalent:
        deal.send_equivalent(settings_portal.sum_equivalent_code, float(sum_equivalent))

    return JsonResponse({'result': 'ok'})


@xframe_options_exempt
@csrf_exempt
def export_excel(request):
    """Метод экспорта в excel."""

    member_id = request.GET.get('member_id')
    quote_id = int(request.GET.get('quote_id'))
    portal: Portals = create_portal(member_id)

    products = ProdTimeQuote.objects.filter(portal=portal, quote_id=quote_id)
    products = sorted(products, key=lambda prod: prod.sort)

    quote = QuoteB24(portal, quote_id)

    try:
        company = CompanyB24(portal, quote.company_id)
        company_name = company.name
    except RuntimeError:
        company_name = 'Не удалось получить наименование компании'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.'
                     'spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
        quote.properties.get('UF_CRM_630DB207D313F') or 'report'
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Prodtime'

    columns = [
        'Товар',
        'Кол-во',
        'Кол-во раб. дней',
        'Срок производства'
    ]

    worksheet.merge_cells('B1:D1')
    worksheet.merge_cells('B2:D2')
    worksheet.cell(1, 1).value = 'ИД предложения:'
    worksheet.cell(1, 2).value = quote_id
    worksheet.cell(1, 2).alignment = Alignment(horizontal='left')
    worksheet.cell(2, 1).value = 'Компания предложения:'
    worksheet.cell(2, 2).value = company_name

    row_num = 4
    worksheet.column_dimensions['A'].width = 70
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 17
    worksheet.column_dimensions['D'].width = 20

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for product in products:
        row_num += 1
        prod_time = (product.prod_time.strftime('%d.%m.%Y') if
                     product.prod_time else '')

        row = [
            product.name_for_print,
            product.quantity,
            product.count_days,
            prod_time,
        ]

        for col_num, cell_value in enumerate(row, 1):
            worksheet.row_dimensions[row_num].height = 30
            cell = worksheet.cell(row=row_num, column=col_num)
            alignment = Alignment(vertical='center', wrap_text=True,
                                  horizontal='left')
            cell.alignment = alignment
            cell.value = cell_value

    workbook.save(response)

    return response
